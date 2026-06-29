"""
任务一：构建结构化财报数据库
=================================
功能：
1. 扫描所有PDF财报文件
2. 提取文本和表格数据
3. 规则提取 + LLM增强提取财务数据
4. 数据校验和清洗
5. 存入SQLite结构化数据库
6. 导入公司基本信息
7. 打印详细处理进度

使用方法：
    python task1/run_task1.py
"""
import sys
import os
import asyncio
import json
import logging
import time
import concurrent.futures
from pathlib import Path
from datetime import datetime

# 添加项目根目录到路径
PROJECT_ROOT = Path(__file__).parent.parent.resolve()
sys.path.insert(0, str(PROJECT_ROOT))

from backend.core.config import AppConfig, SAMPLE_DATA_DIR, DATA_DIR, RESULTS_DIR
from backend.core.database import DatabaseManager
from backend.core.llm_client import LLMClient
from backend.core.pdf_parser import (
    scan_report_files, extract_text_from_pdf, extract_tables_from_pdf,
    extract_text_and_tables_from_pdf,
    classify_report_by_content, extract_financial_data_by_rules,
    ReportMeta, FinancialData,
)

# =============================================================================
# 日志配置
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(str(PROJECT_ROOT / "logs" / "task1.log"), encoding="utf-8"),
    ]
)
logger = logging.getLogger(__name__)


# =============================================================================
# 公司信息导入
# =============================================================================
def import_company_info(db: DatabaseManager, data_dir: Path):
    """导入附件1的公司基本信息"""
    import openpyxl
    
    xlsx_files = []
    for f in os.listdir(data_dir):
        if f.endswith('.xlsx') and ('1' in f or '基本信息' in f or '公司' in f):
            xlsx_files.append(data_dir / f)
    
    if not xlsx_files:
        # 尝试找到附件1
        for f in os.listdir(data_dir):
            fp = data_dir / f
            if f.endswith('.xlsx') and fp.is_file():
                try:
                    wb = openpyxl.load_workbook(str(fp))
                    if len(wb.sheetnames) == 2:  # 附件1有2个sheet
                        xlsx_files.append(fp)
                        break
                except Exception:
                    pass
    
    if not xlsx_files:
        logger.warning("未找到公司基本信息文件(附件1)")
        return
    
    xlsx_path = str(xlsx_files[0])
    logger.info(f"导入公司基本信息: {xlsx_path}")
    
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]  # 第一个sheet是基本信息表
    
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        logger.warning("公司信息表数据为空")
        return
    
    headers = [str(h) for h in rows[0]]
    records = []
    
    for row in rows[1:]:
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = val
        
        # 映射到数据库字段
        # 股票代码补全前导零至6位（Excel中000999会被读成整数999）
        raw_code = str(record.get("股票代码", ""))
        stock_code = raw_code.zfill(6) if raw_code and raw_code.isdigit() else raw_code
        db_record = {
            "serial_number": record.get("序号"),
            "stock_code": stock_code,
            "stock_abbr": record.get("A股简称", ""),
            "company_name": record.get("公司名称", ""),
            "english_name": record.get("英文名称", ""),
            "industry": record.get("所属证监会行业", ""),
            "exchange": record.get("上市交易所", ""),
            "security_type": record.get("证券类别", ""),
            "registered_area": record.get("注册区域", ""),
            "registered_capital": str(record.get("注册资本", "")),
            "employee_count": record.get("雇员人数"),
            "management_count": record.get("管理人员人数"),
        }
        records.append(db_record)
        logger.info(f"  公司: {db_record['stock_abbr']} ({db_record['stock_code']})")
    
    db.insert_many("company_info", records)
    logger.info(f"成功导入 {len(records)} 家公司信息")
    return records


# =============================================================================
# 构建股票代码→简称映射
# =============================================================================
def build_stock_mapping(db: DatabaseManager) -> dict:
    """从数据库获取股票代码到简称的映射"""
    rows = db.execute_query("SELECT stock_code, stock_abbr FROM company_info")
    mapping = {}
    for row in rows:
        mapping[str(row['stock_code'])] = row['stock_abbr']
    return mapping


# =============================================================================
# LLM增强提取
# =============================================================================
LLM_EXTRACT_PROMPT = """从以下财报文本提取财务数据，返回纯JSON（不要markdown代码块）。
公司:{stock_abbr}({stock_code}) 报告期:{report_period} 类型:{report_type}

提取规则:
- 金额单位统一为万元。如原文单位是"元"则÷10000
- 找不到的字段设为null
- 百分比直接填数字(如15.5表示15.5%)
- eps/net_asset_per_share/operating_cf_per_share单位为元

字段说明:
income_sheet: total_operating_revenue(营业总收入), operating_expense_cost_of_sales(营业成本), operating_expense_selling_expenses(销售费用), operating_expense_administrative_expenses(管理费用), operating_expense_financial_expenses(财务费用), operating_expense_rnd_expenses(研发费用), operating_expense_taxes_and_surcharges(税金及附加), total_operating_expenses(营业总成本), operating_profit(营业利润), total_profit(利润总额), net_profit(净利润), other_income(其他收益), investment_income(投资收益), fair_value_change_income(公允价值变动收益), asset_disposal_income(资产处置收益), non_operating_income(营业外收入), non_operating_expenses(营业外支出), asset_impairment_loss(资产减值损失), credit_impairment_loss(信用减值损失)

balance_sheet: asset_cash_and_cash_equivalents(货币资金), asset_accounts_receivable(应收账款), asset_inventory(存货), asset_trading_financial_assets(交易性金融资产), asset_construction_in_progress(在建工程), asset_total_assets(资产总计), liability_accounts_payable(应付账款), liability_advance_from_customers(预收账款), liability_total_liabilities(负债合计), liability_contract_liabilities(合同负债), liability_short_term_loans(短期借款), equity_unappropriated_profit(未分配利润), equity_total_equity(股东权益合计)

cash_flow_sheet: operating_cf_net_amount(经营活动现金流量净额), operating_cf_cash_from_sales(销售商品收到的现金), investing_cf_net_amount(投资活动现金流量净额), investing_cf_cash_for_investments(投资支付的现金), investing_cf_cash_from_investment_recovery(收回投资收到的现金), financing_cf_cash_from_borrowing(取得借款收到的现金), financing_cf_cash_for_debt_repayment(偿还债务支付的现金), financing_cf_net_amount(筹资活动现金流量净额), net_cash_flow(现金及现金等价物净增加额)

core_performance: eps(基本每股收益/元), net_asset_per_share(每股净资产/元), roe(净资产收益率/%), operating_cf_per_share(每股经营现金流量/元), gross_profit_margin(销售毛利率/%)

财报文本:
{text_content}

返回JSON:
{{"income_sheet":{{...}},"balance_sheet":{{...}},"cash_flow_sheet":{{...}},"core_performance":{{...}}}}"""


async def llm_enhance_extraction(
    llm: LLMClient,
    text: str,
    meta: ReportMeta,
    rule_data: FinancialData,
) -> FinancialData:
    """使用LLM增强数据提取"""
    text_truncated = text[:20000]
    
    prompt = LLM_EXTRACT_PROMPT.format(
        stock_abbr=meta.stock_abbr,
        stock_code=meta.stock_code,
        report_period=meta.report_period,
        report_type=meta.report_type,
        text_content=text_truncated,
    )
    
    try:
        result = await asyncio.wait_for(
            llm.query_json(prompt, temperature=0.1, prefer_agent=False),
            timeout=60
        )
        
        # 合并结果：LLM填补规则提取的空值，也可覆盖规则值
        for sheet_key in ["income_sheet", "balance_sheet", "cash_flow_sheet", "core_performance"]:
            llm_data = result.get(sheet_key, {})
            if not isinstance(llm_data, dict):
                continue
            rule_sheet = getattr(rule_data, sheet_key, {}) or {}
            
            for k, v in llm_data.items():
                if v is None:
                    continue
                # 尝试转为数值
                if isinstance(v, str):
                    v = v.replace(",", "").replace("，", "").strip()
                    try:
                        v = float(v)
                    except (ValueError, TypeError):
                        continue
                # 填补空值 或 覆盖可能有误的规则值
                if k not in rule_sheet or rule_sheet[k] is None:
                    rule_sheet[k] = v
            
            setattr(rule_data, sheet_key, rule_sheet)
        
        logger.info(f"  LLM增强提取完成: {meta.file_name}")
    except Exception as e:
        logger.warning(f"  LLM增强提取失败: {e}")
    
    return rule_data


# =============================================================================
# 数据校验
# =============================================================================
def validate_financial_data(data: FinancialData) -> list:
    """
    多维度数据校验
    返回校验问题列表
    """
    issues = []
    meta = data.meta
    prefix = f"[{meta.stock_abbr} {meta.report_period}]"
    
    income = data.income_sheet
    balance = data.balance_sheet
    cash_flow = data.cash_flow_sheet
    
    # 1. 利润表校验：使用中国会计准则正确公式
    # 营业利润 = 营业总收入 - 营业总成本 + 其他收益 + 投资收益 + 公允价值变动收益
    #            + 信用减值损失(已含符号) + 资产减值损失(已含符号) + 资产处置收益
    rev = income.get("total_operating_revenue")
    exp = income.get("total_operating_expenses")
    op_profit = income.get("operating_profit")
    if rev is not None and exp is not None and op_profit is not None:
        calc_profit = rev - exp
        # 加上已提取的中间项（中国会计准则中这些项目介于营业总成本和营业利润之间）
        other_items = {
            "other_income": income.get("other_income"),
            "investment_income": income.get("investment_income"),
            "fair_value_change_income": income.get("fair_value_change_income"),
            "credit_impairment_loss": income.get("credit_impairment_loss"),
            "asset_impairment_loss": income.get("asset_impairment_loss"),
            "asset_disposal_income": income.get("asset_disposal_income"),
        }
        extracted_count = 0
        for k, v in other_items.items():
            if v is not None:
                calc_profit += v
                extracted_count += 1
        
        diff = abs(calc_profit - op_profit)
        total_items = len(other_items)  # 6个中间项
        missing_count = total_items - extracted_count
        
        # 根据中间项覆盖率动态调整容差
        # 缺少的中间项越多，容差越大；全部缺失时跳过校验
        if extracted_count == 0:
            threshold = None  # 无中间项时跳过校验（公式不完整，无法可靠校验）
        elif missing_count == 0:
            threshold = abs(op_profit) * 0.05 + 100  # 全部提取到，严格校验
        elif missing_count <= 2:
            threshold = abs(op_profit) * 0.5 + abs(rev) * 0.15 + 5000  # 缺少1-2项，放宽（缺失项可能数千万）
        else:
            threshold = abs(op_profit) * 1.0 + abs(rev) * 0.2 + 10000  # 缺少3+项，大幅放宽
        
        if threshold is not None and diff > threshold:
            extracted_items = {k: v for k, v in other_items.items() if v is not None}
            issues.append(
                f"{prefix} 利润表校验({extracted_count}/{total_items}项): "
                f"营业总收入({rev})-营业总成本({exp})+中间项={calc_profit}, "
                f"但营业利润={op_profit}, 差额={diff}"
            )
    
    # 1b. 利润总额校验：利润总额 = 营业利润 + 营业外收入 - 营业外支出
    total_profit = income.get("total_profit")
    non_op_income = income.get("non_operating_income")
    non_op_expense = income.get("non_operating_expenses")
    if op_profit is not None and total_profit is not None:
        if non_op_income is not None and non_op_expense is not None:
            calc_total = op_profit + non_op_income - non_op_expense
            diff = abs(calc_total - total_profit)
            if diff > abs(total_profit) * 0.05 + 50:
                issues.append(
                    f"{prefix} 利润总额校验: 营业利润({op_profit})+营业外收入({non_op_income})"
                    f"-营业外支出({non_op_expense})={calc_total}, 但利润总额={total_profit}, 差额={diff}"
                )
    
    # 2. 资产负债表校验：总资产 = 总负债 + 股东权益
    total_assets = balance.get("asset_total_assets")
    total_liab = balance.get("liability_total_liabilities")
    total_equity = balance.get("equity_total_equity")
    if total_assets is not None and total_liab is not None and total_equity is not None:
        calc_assets = total_liab + total_equity
        diff = abs(calc_assets - total_assets)
        if diff > abs(total_assets) * 0.01 + 10:  # 允许1%误差或10万元
            issues.append(
                f"{prefix} 资产负债表校验: 总负债({total_liab})+股东权益({total_equity})={calc_assets}, "
                f"但总资产={total_assets}, 差额={diff}"
            )
    
    # 3. 资产负债率校验
    if total_assets is not None and total_liab is not None and total_assets != 0:
        calc_ratio = total_liab / total_assets * 100
        reported_ratio = balance.get("asset_liability_ratio")
        if reported_ratio is not None:
            diff = abs(calc_ratio - reported_ratio)
            if diff > 1:  # 允许1个百分点误差
                issues.append(
                    f"{prefix} 资产负债率校验: 计算值={calc_ratio:.2f}%, 报告值={reported_ratio}%"
                )
    
    # 4. 数据完整性检查
    required_income_fields = ["total_operating_revenue", "net_profit", "total_profit"]
    for field in required_income_fields:
        if income.get(field) is None:
            issues.append(f"{prefix} 利润表缺失关键字段: {field}")
    
    required_balance_fields = ["asset_total_assets", "liability_total_liabilities", "equity_total_equity"]
    for field in required_balance_fields:
        if balance.get(field) is None:
            issues.append(f"{prefix} 资产负债表缺失关键字段: {field}")
    
    # 5. 数值合理性检查
    if income.get("net_profit") is not None and income.get("total_operating_revenue") is not None:
        if income["total_operating_revenue"] != 0:
            margin = income["net_profit"] / income["total_operating_revenue"] * 100
            if abs(margin) > 100:
                issues.append(
                    f"{prefix} 净利率异常: {margin:.1f}% (净利润/营业收入)"
                )
    
    return issues


async def llm_validate_data(llm: LLMClient, data: FinancialData, issues: list) -> list:
    """使用LLM辅助校验"""
    if not issues:
        return issues
    
    prompt = f"""
以下是从财务报告中提取的数据校验结果，请分析这些问题并给出建议：

公司: {data.meta.stock_abbr} ({data.meta.stock_code})
报告期: {data.meta.report_period}

校验发现的问题:
{json.dumps(issues, ensure_ascii=False, indent=2)}

提取的利润表数据:
{json.dumps(data.income_sheet, ensure_ascii=False, indent=2)}

提取的资产负债表数据:
{json.dumps(data.balance_sheet, ensure_ascii=False, indent=2)}

请分析：
1. 哪些问题是真正的数据错误？
2. 哪些可能是因为报告格式导致的提取误差？
3. 建议如何修正？

返回JSON格式：
{{
    "real_issues": ["真正需要关注的问题"],
    "corrections": {{"字段名": "建议修正值"}},
    "notes": "其他说明"
}}
"""
    try:
        result = await llm.query_json(prompt, temperature=0.1, prefer_agent=False)
        return result.get("real_issues", issues)
    except Exception:
        return issues


# =============================================================================
# 数据入库
# =============================================================================
_TABLE_COLUMNS_CACHE = {}

def _get_valid_columns(db: DatabaseManager, table: str) -> set:
    """获取表的有效列名（带缓存）"""
    if table not in _TABLE_COLUMNS_CACHE:
        info = db.get_table_info(table)
        _TABLE_COLUMNS_CACHE[table] = {row["name"] for row in info}
    return _TABLE_COLUMNS_CACHE[table]


def _sanitize_record(db: DatabaseManager, table: str, record: dict) -> dict:
    """过滤掉不属于该表的字段，避免SQL错误"""
    valid = _get_valid_columns(db, table)
    return {k: v for k, v in record.items() if k in valid and "/" not in k}


def save_to_database(db: DatabaseManager, data: FinancialData):
    """将提取的财务数据存入数据库"""
    meta = data.meta
    
    common_fields = {
        "stock_code": meta.stock_code,
        "stock_abbr": meta.stock_abbr,
        "report_period": meta.report_period,
        "report_year": meta.report_year,
    }
    
    # 利润表
    if data.income_sheet:
        record = {**common_fields, **data.income_sheet}
        record.pop("serial_number", None)
        db.insert_record("income_sheet", _sanitize_record(db, "income_sheet", record))
    
    # 资产负债表
    if data.balance_sheet:
        record = {**common_fields, **data.balance_sheet}
        record.pop("serial_number", None)
        db.insert_record("balance_sheet", _sanitize_record(db, "balance_sheet", record))
    
    # 现金流量表
    if data.cash_flow_sheet:
        record = {**common_fields, **data.cash_flow_sheet}
        record.pop("serial_number", None)
        db.insert_record("cash_flow_sheet", _sanitize_record(db, "cash_flow_sheet", record))
    
    # 核心业绩指标
    if data.core_performance:
        perf = {**common_fields, **data.core_performance}
        if "total_operating_revenue" not in perf and "total_operating_revenue" in data.income_sheet:
            perf["total_operating_revenue"] = data.income_sheet["total_operating_revenue"]
        if "net_profit_10k_yuan" not in perf and "net_profit" in data.income_sheet:
            perf["net_profit_10k_yuan"] = data.income_sheet["net_profit"]
        perf.pop("serial_number", None)
        db.insert_record("core_performance_indicators_sheet", _sanitize_record(db, "core_performance_indicators_sheet", perf))
    
    logger.info(f"  数据已入库: {meta.stock_abbr} {meta.report_period}")


# =============================================================================
# 计算同比/环比增长率
# =============================================================================
def _safe_float(val):
    """安全转换为float，处理LLM返回的字符串值"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(",", "").replace("，", "").strip())
    except (ValueError, TypeError):
        return None


def calculate_growth_rates(db: DatabaseManager):
    """在所有数据入库后，计算同比和环比增长率"""
    logger.info("\n" + "=" * 60)
    logger.info("计算同比和环比增长率...")
    
    # 获取所有公司和报告期
    periods = db.execute_query("""
        SELECT DISTINCT stock_code, stock_abbr, report_period, report_year 
        FROM income_sheet ORDER BY stock_code, report_period
    """)
    
    for p in periods:
        code = p['stock_code']
        period = p['report_period']
        year = p['report_year']
        
        # 确定同比对应期（去年同期）
        suffix = period.replace(str(year), "")
        yoy_period = f"{year - 1}{suffix}"
        
        # 利润表同比
        current = db.execute_query(
            "SELECT * FROM income_sheet WHERE stock_code=? AND report_period=?", (code, period)
        )
        prev = db.execute_query(
            "SELECT * FROM income_sheet WHERE stock_code=? AND report_period=?", (code, yoy_period)
        )
        
        if current and prev:
            curr_rev = _safe_float(current[0].get("total_operating_revenue"))
            prev_rev = _safe_float(prev[0].get("total_operating_revenue"))
            curr_profit = _safe_float(current[0].get("net_profit"))
            prev_profit = _safe_float(prev[0].get("net_profit"))
            
            updates = {}
            if curr_rev is not None and prev_rev is not None and prev_rev != 0:
                updates["operating_revenue_yoy_growth"] = round((curr_rev - prev_rev) / abs(prev_rev) * 100, 4)
            if curr_profit is not None and prev_profit is not None and prev_profit != 0:
                updates["net_profit_yoy_growth"] = round((curr_profit - prev_profit) / abs(prev_profit) * 100, 4)
            
            if updates:
                set_clause = ", ".join([f"{k}=?" for k in updates.keys()])
                db.execute_sql(
                    f"UPDATE income_sheet SET {set_clause} WHERE stock_code=? AND report_period=?",
                    tuple(updates.values()) + (code, period)
                )
                
                # 同步更新核心业绩表
                for k, v in updates.items():
                    if k in ["operating_revenue_yoy_growth", "net_profit_yoy_growth"]:
                        db.execute_sql(
                            f"UPDATE core_performance_indicators_sheet SET {k}=? WHERE stock_code=? AND report_period=?",
                            (v, code, period)
                        )
        
        # 资产负债表同比
        current_bs = db.execute_query(
            "SELECT * FROM balance_sheet WHERE stock_code=? AND report_period=?", (code, period)
        )
        prev_bs = db.execute_query(
            "SELECT * FROM balance_sheet WHERE stock_code=? AND report_period=?", (code, yoy_period)
        )
        
        if current_bs and prev_bs:
            updates = {}
            for field, growth_field in [
                ("asset_total_assets", "asset_total_assets_yoy_growth"),
                ("liability_total_liabilities", "liability_total_liabilities_yoy_growth"),
            ]:
                curr_val = _safe_float(current_bs[0].get(field))
                prev_val = _safe_float(prev_bs[0].get(field))
                if curr_val is not None and prev_val is not None and prev_val != 0:
                    updates[growth_field] = round((curr_val - prev_val) / abs(prev_val) * 100, 4)
            
            if updates:
                set_clause = ", ".join([f"{k}=?" for k in updates.keys()])
                db.execute_sql(
                    f"UPDATE balance_sheet SET {set_clause} WHERE stock_code=? AND report_period=?",
                    tuple(updates.values()) + (code, period)
                )
        
        # 现金流量表同比
        current_cf = db.execute_query(
            "SELECT * FROM cash_flow_sheet WHERE stock_code=? AND report_period=?", (code, period)
        )
        prev_cf = db.execute_query(
            "SELECT * FROM cash_flow_sheet WHERE stock_code=? AND report_period=?", (code, yoy_period)
        )
        
        if current_cf and prev_cf:
            curr_ncf = _safe_float(current_cf[0].get("net_cash_flow"))
            prev_ncf = _safe_float(prev_cf[0].get("net_cash_flow"))
            if curr_ncf is not None and prev_ncf is not None and prev_ncf != 0:
                growth = round((curr_ncf - prev_ncf) / abs(prev_ncf) * 100, 4)
                db.execute_sql(
                    "UPDATE cash_flow_sheet SET net_cash_flow_yoy_growth=? WHERE stock_code=? AND report_period=?",
                    (growth, code, period)
                )
    
    logger.info("同比/环比增长率计算完成")


def compute_derived_metrics(db: DatabaseManager):
    """从已入库的数据中计算衍生指标，填补空缺字段"""
    logger.info("计算衍生指标...")

    # 1. gross_profit_margin / net_profit_margin from income_sheet
    rows = db.execute_query("SELECT stock_code, report_period FROM income_sheet")
    for r in rows:
        code, period = r['stock_code'], r['report_period']
        inc = db.execute_query(
            "SELECT total_operating_revenue, operating_expense_cost_of_sales, net_profit "
            "FROM income_sheet WHERE stock_code=? AND report_period=?", (code, period)
        )
        if not inc:
            continue
        rev = _safe_float(inc[0].get("total_operating_revenue"))
        cost = _safe_float(inc[0].get("operating_expense_cost_of_sales"))
        np_ = _safe_float(inc[0].get("net_profit"))

        updates_core = {}
        if rev and rev != 0 and cost is not None:
            updates_core["gross_profit_margin"] = round((rev - cost) / rev * 100, 4)
        if rev and rev != 0 and np_ is not None:
            updates_core["net_profit_margin"] = round(np_ / rev * 100, 4)

        if updates_core:
            # 仅填充空值
            existing = db.execute_query(
                "SELECT gross_profit_margin, net_profit_margin FROM core_performance_indicators_sheet "
                "WHERE stock_code=? AND report_period=?", (code, period)
            )
            if existing:
                for k in list(updates_core.keys()):
                    if existing[0].get(k) is not None:
                        del updates_core[k]
                if updates_core:
                    set_clause = ", ".join([f"{k}=?" for k in updates_core])
                    db.execute_sql(
                        f"UPDATE core_performance_indicators_sheet SET {set_clause} "
                        f"WHERE stock_code=? AND report_period=?",
                        tuple(updates_core.values()) + (code, period)
                    )

    # 2. QoQ growth (环比)
    periods_all = db.execute_query(
        "SELECT DISTINCT stock_code, report_period, report_year FROM income_sheet ORDER BY stock_code, report_period"
    )
    qoq_map = {"Q1": None, "HY": "Q1", "Q3": "HY", "FY": "Q3"}
    for p in periods_all:
        code, period, year = p['stock_code'], p['report_period'], p['report_year']
        if not year:
            continue
        suffix = period.replace(str(year), "")
        prev_suffix = qoq_map.get(suffix)
        if prev_suffix is None:
            continue
        prev_period = f"{year}{prev_suffix}"

        cur = db.execute_query(
            "SELECT total_operating_revenue, net_profit FROM income_sheet WHERE stock_code=? AND report_period=?",
            (code, period)
        )
        prev = db.execute_query(
            "SELECT total_operating_revenue, net_profit FROM income_sheet WHERE stock_code=? AND report_period=?",
            (code, prev_period)
        )
        if not cur or not prev:
            continue

        qoq_updates = {}
        cr = _safe_float(cur[0].get("total_operating_revenue"))
        pr = _safe_float(prev[0].get("total_operating_revenue"))
        if cr is not None and pr is not None and pr != 0:
            qoq_updates["operating_revenue_qoq_growth"] = round((cr - pr) / abs(pr) * 100, 4)
        cn = _safe_float(cur[0].get("net_profit"))
        pn = _safe_float(prev[0].get("net_profit"))
        if cn is not None and pn is not None and pn != 0:
            qoq_updates["net_profit_qoq_growth"] = round((cn - pn) / abs(pn) * 100, 4)

        if qoq_updates:
            set_clause = ", ".join([f"{k}=?" for k in qoq_updates])
            db.execute_sql(
                f"UPDATE core_performance_indicators_sheet SET {set_clause} "
                f"WHERE stock_code=? AND report_period=?",
                tuple(qoq_updates.values()) + (code, period)
            )

    # 3. net_profit_excl_non_recurring YoY
    npe_rows = db.execute_query(
        "SELECT stock_code, report_period, report_year, net_profit_excl_non_recurring "
        "FROM core_performance_indicators_sheet WHERE net_profit_excl_non_recurring IS NOT NULL"
    )
    for r in npe_rows:
        code, period, year = r['stock_code'], r['report_period'], r['report_year']
        if not year:
            continue
        suffix = period.replace(str(year), "")
        yoy_period = f"{year - 1}{suffix}"
        prev_r = db.execute_query(
            "SELECT net_profit_excl_non_recurring FROM core_performance_indicators_sheet "
            "WHERE stock_code=? AND report_period=?", (code, yoy_period)
        )
        if prev_r:
            cv = _safe_float(r['net_profit_excl_non_recurring'])
            pv = _safe_float(prev_r[0].get('net_profit_excl_non_recurring'))
            if cv is not None and pv is not None and pv != 0:
                growth = round((cv - pv) / abs(pv) * 100, 4)
                db.execute_sql(
                    "UPDATE core_performance_indicators_sheet SET net_profit_excl_non_recurring_yoy=? "
                    "WHERE stock_code=? AND report_period=?",
                    (growth, code, period)
                )

    logger.info("衍生指标计算完成")


# =============================================================================
# 主流程
# =============================================================================
def _process_one_pdf(args):
    """Phase 1 工作函数：在进程/线程池中运行，纯本地，无API调用"""
    idx, meta, stock_mapping = args
    try:
        text, tables = extract_text_and_tables_from_pdf(meta.file_path)
        meta = classify_report_by_content(text, meta)

        # 双向补齐：代码→简称 和 简称→代码
        if not meta.stock_abbr and meta.stock_code in stock_mapping:
            meta.stock_abbr = stock_mapping[meta.stock_code]
        if not meta.stock_code and meta.stock_abbr:
            for code, abbr in stock_mapping.items():
                if abbr == meta.stock_abbr:
                    meta.stock_code = code
                    break

        if "摘要" in meta.report_type:
            return {"status": "skip", "idx": idx}

        if not meta.report_period:
            return {"status": "fail", "idx": idx, "error": "无法确定报告期"}

        data = extract_financial_data_by_rules(text, tables, meta)
        rule_fields = sum(
            len(v) for v in [data.income_sheet, data.balance_sheet,
                             data.cash_flow_sheet, data.core_performance] if v
        )

        return {
            "status": "ok",
            "idx": idx,
            "data": data,
            "meta": meta,
            "text": text[:20000],  # 保留截断文本供Phase 2 LLM使用
            "rule_fields": rule_fields,
        }
    except Exception as e:
        return {"status": "fail", "idx": idx, "error": str(e)}


async def main():
    start_time = time.time()

    print("=" * 70)
    print("  任务一：构建结构化财报数据库（两阶段优化版）")
    print("  开始时间:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    print("=" * 70)

    # 1. 加载配置
    config = AppConfig.load()
    print(f"\n[1/9] 加载配置完成")
    print(f"  LLM API: {[c.name for c in config.get_enabled_llms()]}")
    print(f"  数据目录: {SAMPLE_DATA_DIR}")

    # 2. 初始化数据库（清空旧数据重新构建）
    db = DatabaseManager(config.db_path)
    db.init_db()
    for _tbl in ["income_sheet", "balance_sheet", "cash_flow_sheet", "core_performance_indicators_sheet"]:
        try:
            db.execute_sql(f"DELETE FROM {_tbl}")
        except Exception:
            pass
    print(f"\n[2/9] 数据库初始化完成(已清空旧财务数据): {config.db_path}")

    # 3. 导入公司基本信息
    print(f"\n[3/9] 导入公司基本信息...")
    stock_mapping = {}
    try:
        records = import_company_info(db, SAMPLE_DATA_DIR)
        if records:
            for r in records:
                stock_mapping[str(r['stock_code'])] = r['stock_abbr']
                print(f"  ✓ {r['stock_abbr']} (代码: {r['stock_code']})")
    except Exception as e:
        logger.error(f"导入公司信息失败: {e}")

    if not stock_mapping:
        stock_mapping = build_stock_mapping(db)

    # 4. 扫描财报文件
    print(f"\n[4/9] 扫描财报文件...")
    reports = scan_report_files(str(SAMPLE_DATA_DIR))
    total = len(reports)
    print(f"  共发现 {total} 个PDF文件")

    # 建立股票代码与简称的双向映射，确保同一公司的报告被分到同一组
    code_to_abbr = stock_mapping.copy()
    abbr_to_code = {v: k for k, v in stock_mapping.items()}
    
    by_company = {}
    for r in reports:
        # 优先使用股票代码作为分组键
        if r.stock_code:
            key = r.stock_code
        elif r.stock_abbr and r.stock_abbr in abbr_to_code:
            # 如果只有简称，通过反向映射查找股票代码
            key = abbr_to_code[r.stock_abbr]
        elif r.stock_abbr:
            # 简称不在映射中，直接使用简称
            key = r.stock_abbr
        else:
            key = "未知"
        by_company.setdefault(key, []).append(r)
    
    # 打印统计信息（优先显示简称）
    for company, reps in sorted(by_company.items()):
        display_name = code_to_abbr.get(company, company)
        print(f"  - {display_name} ({company}): {len(reps)} 份报告")

    # ================================================================
    # Phase 1: 纯本地规则提取（线程池并行，无API调用）
    # ================================================================
    print(f"\n[5/9] ===== Phase 1: 规则提取所有PDF（纯本地，无API）=====")
    phase1_start = time.time()

    args_list = [(i, meta, stock_mapping) for i, meta in enumerate(reports)]

    results = []
    ok_count = 0
    skip_count = 0
    fail_count = 0

    NUM_WORKERS = min(8, os.cpu_count() or 4)
    print(f"  进程池大小: {NUM_WORKERS}", flush=True)

    with concurrent.futures.ProcessPoolExecutor(max_workers=NUM_WORKERS) as executor:
        futures = {executor.submit(_process_one_pdf, a): a[0] for a in args_list}
        done_count = 0
        for future in concurrent.futures.as_completed(futures):
            done_count += 1
            r = future.result()
            results.append(r)

            if r["status"] == "ok":
                ok_count += 1
            elif r["status"] == "skip":
                skip_count += 1
            else:
                fail_count += 1
                logger.error(f"  [{r['idx']+1}/{total}] 失败: {r.get('error')}")

            if done_count % 100 == 0 or done_count == total:
                elapsed = time.time() - phase1_start
                rate = done_count / elapsed if elapsed > 0 else 0
                eta = (total - done_count) / rate if rate > 0 else 0
                print(f"  Phase1 进度: {done_count}/{total} ({done_count*100//total}%) | "
                      f"成功{ok_count} 跳过{skip_count} 失败{fail_count} | "
                      f"{rate:.1f}个/秒 | 剩余~{eta:.0f}秒", flush=True)

    phase1_time = time.time() - phase1_start
    print(f"\n  Phase 1 完成! 耗时 {phase1_time:.1f}秒", flush=True)
    print(f"  成功提取: {ok_count}, 跳过摘要: {skip_count}, 失败: {fail_count}", flush=True)

    # ================================================================
    # Phase 1.5: 入库（规则提取结果先全部入库）
    # ================================================================
    print(f"\n[6/9] 规则提取数据入库...")
    ok_results = [r for r in results if r["status"] == "ok"]
    need_llm_results = []
    all_issues = []

    for r in ok_results:
        data = r["data"]
        try:
            issues = validate_financial_data(data)
            if issues:
                all_issues.extend(issues)
            save_to_database(db, data)
        except Exception as e:
            logger.error(f"  入库失败 [{r['idx']+1}]: {e}")

        # 所有报告都送LLM增强（最大化数据完整性）
        need_llm_results.append(r)

    print(f"  已入库 {len(ok_results)} 条记录")
    print(f"  需要LLM增强: {len(need_llm_results)} 条 (全部送LLM填补空缺)")

    # ================================================================
    # Phase 2: 批量LLM增强（仅对不足的报告）
    # ================================================================
    print(f"\n[7/9] ===== Phase 2: LLM批量增强 ({len(need_llm_results)}条) =====")
    phase2_start = time.time()

    if need_llm_results:
        llm = LLMClient(config)
        print(f"  LLM API: {[c.name for c in config.get_enabled_llms()]}")

        llm_done = 0
        llm_ok = 0
        llm_fail = 0
        LLM_CONCURRENCY = 30
        sem = asyncio.Semaphore(LLM_CONCURRENCY)

        async def _llm_enhance_one(r):
            nonlocal llm_done, llm_ok, llm_fail
            async with sem:
                try:
                    data = r["data"]
                    meta = r["meta"]
                    text = r["text"]
                    data = await llm_enhance_extraction(llm, text, meta, data)

                    # 更新入库（REPLACE模式会覆盖旧数据）
                    save_to_database(db, data)
                    llm_ok += 1
                except Exception as e:
                    llm_fail += 1
                    logger.error(f"  LLM增强失败 [{r['idx']+1}]: {e}")
                finally:
                    llm_done += 1
                    if llm_done % 100 == 0 or llm_done == len(need_llm_results):
                        elapsed = time.time() - phase2_start
                        rate = llm_done / elapsed if elapsed > 0 else 0
                        eta = (len(need_llm_results) - llm_done) / rate if rate > 0 else 0
                        print(f"  Phase2 进度: {llm_done}/{len(need_llm_results)} | "
                              f"成功{llm_ok} 失败{llm_fail} | "
                              f"{rate:.1f}个/秒 | 剩余~{eta:.0f}秒", flush=True)

        tasks = [_llm_enhance_one(r) for r in need_llm_results]
        await asyncio.gather(*tasks)

        phase2_time = time.time() - phase2_start
        print(f"\n  Phase 2 完成! 耗时 {phase2_time:.1f}秒, 成功{llm_ok}, 失败{llm_fail}")
    else:
        print("  无需LLM增强，跳过")

    # 8. 计算增长率
    print(f"\n[8/10] 计算同比/环比增长率...")
    calculate_growth_rates(db)

    # 9. 计算衍生指标
    print(f"\n[9/10] 计算衍生指标(毛利率/净利率/环比等)...")
    compute_derived_metrics(db)

    # 10. 汇总报告
    elapsed = time.time() - start_time
    print(f"\n[9/9] 处理完成!")
    print("=" * 70)
    print(f"  总计文件: {total}")
    print(f"  规则提取成功: {ok_count}")
    print(f"  跳过摘要: {skip_count}")
    print(f"  失败: {fail_count}")
    print(f"  LLM增强: {len(need_llm_results)} 条")
    print(f"  校验问题: {len(all_issues)}")
    print(f"  总耗时: {elapsed:.1f} 秒")
    print()

    print("数据库统计:")
    for table in ["company_info", "core_performance_indicators_sheet", "balance_sheet", "income_sheet", "cash_flow_sheet"]:
        count = db.get_table_row_count(table)
        print(f"  {table}: {count} 条记录")

    if all_issues:
        report_path = str(PROJECT_ROOT / "logs" / "validation_report.txt")
        with open(report_path, "w", encoding="utf-8") as f:
            f.write("数据校验报告\n")
            f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 60 + "\n\n")
            for issue in all_issues:
                f.write(f"- {issue}\n")
        print(f"\n校验报告已保存: {report_path}")

    print("\n" + "=" * 70)
    print("  任务一完成！结构化财报数据库已构建。")
    print("=" * 70)


if __name__ == "__main__":
    # 确保日志目录存在
    os.makedirs(str(PROJECT_ROOT / "logs"), exist_ok=True)
    os.makedirs(str(DATA_DIR), exist_ok=True)
    os.makedirs(str(RESULTS_DIR), exist_ok=True)
    
    asyncio.run(main())
