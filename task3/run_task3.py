"""
任务三：增强"智能问数"助手的可靠性 - 回答附件6的问题
=================================
功能：
1. 加载任务一构建的结构化数据库
2. 构建研报知识库（附件5）
3. 初始化增强版智能问答Agent
4. 处理附件6中的问题（多意图、意图模糊、归因分析）
5. 输出result_3.xlsx提交文件

使用方法：
    python task3/run_task3.py
"""
import sys
import os
import asyncio
import json
import logging
import time
from pathlib import Path
from datetime import datetime

PROJECT_ROOT = Path(__file__).parent.parent.resolve()
sys.path.insert(0, str(PROJECT_ROOT))

from backend.core.config import AppConfig, SAMPLE_DATA_DIR, DATA_DIR, RESULTS_DIR
from backend.core.database import DatabaseManager
from backend.core.llm_client import LLMClient
from backend.core.embedding import EmbeddingManager
from backend.core.knowledge_base import KnowledgeBase
from backend.core.pdf_parser import extract_text_from_pdf
from backend.core.agent import SmartQAAgent

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(str(PROJECT_ROOT / "logs" / "task3.log"), encoding="utf-8"),
    ]
)
logger = logging.getLogger(__name__)


def build_knowledge_base(
    db: DatabaseManager,
    embedding: EmbeddingManager,
    data_dir: Path,
) -> KnowledgeBase:
    """构建研报知识库"""
    kb = KnowledgeBase(db, embedding)
    
    print("\n构建研报知识库...")
    
    # 查找附件5目录
    att5_dir = None
    for d in os.listdir(data_dir):
        full = data_dir / d
        if full.is_dir() and '5' in d:
            att5_dir = full
            break
    
    if not att5_dir:
        logger.warning("未找到附件5目录")
        return kb
    
    print(f"  附件5目录: {att5_dir}")
    
    # 加载研报元数据
    import openpyxl
    
    stock_reports = []
    industry_reports = []
    
    for f in os.listdir(att5_dir):
        if f.endswith('.xlsx'):
            fp = att5_dir / f
            try:
                wb = openpyxl.load_workbook(str(fp))
                for sn in wb.sheetnames:
                    ws = wb[sn]
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) >= 2:
                        headers = [str(h) for h in rows[0]]
                        for row in rows[1:]:
                            record = dict(zip(headers, [str(v) if v else "" for v in row]))
                            if "stockName" in headers:
                                stock_reports.append(record)
                            elif "industryName" in headers:
                                industry_reports.append(record)
            except Exception as e:
                logger.warning(f"  读取 {f} 失败: {e}")
    
    print(f"  个股研报元数据: {len(stock_reports)} 条")
    print(f"  行业研报元数据: {len(industry_reports)} 条")
    
    # 加载个股研报PDF
    stock_report_dir = att5_dir / "个股研报"
    if stock_report_dir.exists():
        for pdf_file in stock_report_dir.glob("*.pdf"):
            print(f"  处理个股研报: {pdf_file.name}")
            try:
                text = extract_text_from_pdf(str(pdf_file))
                
                meta = {}
                for sr in stock_reports:
                    if sr.get("title", "") in pdf_file.name or pdf_file.name.replace(".pdf", "") in sr.get("title", ""):
                        meta = sr
                        break
                
                kb.add_research_report(
                    title=meta.get("title", pdf_file.stem),
                    report_type="individual",
                    file_path=str(pdf_file),
                    content=text,
                    stock_name=meta.get("stockName", ""),
                    stock_code=meta.get("stockCode", ""),
                    org_name=meta.get("orgName", ""),
                    org_sname=meta.get("orgSName", ""),
                    publish_date=meta.get("publishDate", ""),
                    industry_name=meta.get("indvInduName", ""),
                    rating_name=meta.get("emRatingName", ""),
                    researcher=meta.get("researcher", ""),
                )
                print(f"    ✓ 已加入知识库, 文本长度: {len(text)}")
            except Exception as e:
                logger.warning(f"    ✗ 处理失败: {e}")
    
    # 加载行业研报PDF
    industry_report_dir = att5_dir / "行业研报"
    if industry_report_dir.exists():
        for pdf_file in industry_report_dir.glob("*.pdf"):
            print(f"  处理行业研报: {pdf_file.name}")
            try:
                text = extract_text_from_pdf(str(pdf_file))
                
                meta = {}
                for ir in industry_reports:
                    if ir.get("title", "") in pdf_file.name or pdf_file.name.replace(".pdf", "") in ir.get("title", ""):
                        meta = ir
                        break
                
                kb.add_research_report(
                    title=meta.get("title", pdf_file.stem),
                    report_type="industry",
                    file_path=str(pdf_file),
                    content=text,
                    org_name=meta.get("orgName", ""),
                    org_sname=meta.get("orgSName", ""),
                    publish_date=meta.get("publishDate", ""),
                    industry_name=meta.get("industryName", ""),
                    rating_name=meta.get("emRatingName", ""),
                    researcher=meta.get("researcher", ""),
                )
                print(f"    ✓ 已加入知识库, 文本长度: {len(text)}")
            except Exception as e:
                logger.warning(f"    ✗ 处理失败: {e}")
    
    stats = kb.get_stats()
    print(f"\n  知识库统计:")
    print(f"    文档总数: {stats['total_documents']}")
    print(f"    知识块总数: {stats['total_chunks']}")
    print(f"    研报总数: {stats['total_research_reports']}")
    
    return kb


def load_questions(data_dir: Path) -> list:
    """加载附件6的问题"""
    import openpyxl
    
    for f in os.listdir(data_dir):
        if f.endswith('.xlsx') and ('6' in f or '问题' in f):
            fp = data_dir / f
            try:
                wb = openpyxl.load_workbook(str(fp))
                if len(wb.sheetnames) == 1:
                    ws = wb[wb.sheetnames[0]]
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) >= 2 and rows[0][0] and '编号' in str(rows[0][0]):
                        if 'B2' in str(rows[1][0]):
                            questions = []
                            for row in rows[1:]:
                                questions.append({
                                    "id": str(row[0]),
                                    "type": str(row[1]),
                                    "questions": json.loads(str(row[2])),
                                })
                            return questions
            except Exception as e:
                logger.warning(f"读取 {f} 失败: {e}")
    
    logger.warning("未找到附件6的问题文件")
    return []


def _extract_results_from_answers(answers: list) -> dict:
    """
    从agent返回的answers列表中提取SQL、图表类型、图片路径、引用
    answers格式: [{"Q": "...", "A": {"content": "...", "image": [...], "references": [...]}, "sql": "...", "chart_type": "..."}, ...]
    """
    all_sql = []
    all_chart_types = []
    all_images = []
    all_references = []

    for a in answers:
        if not isinstance(a, dict):
            continue
        # SQL在answer的顶层
        if a.get("sql"):
            all_sql.append(a["sql"])
        # chart_type在answer的顶层
        if a.get("chart_type"):
            all_chart_types.append(a["chart_type"])
        # 图片和引用在 A 字典中
        answer_body = a.get("A", {})
        if isinstance(answer_body, dict):
            if answer_body.get("image"):
                for img_path in answer_body["image"]:
                    img_name = os.path.basename(str(img_path))
                    all_images.append(img_name)
            if answer_body.get("references"):
                all_references.extend(answer_body["references"])

    # 合并SQL
    sql_text = ";\n".join(all_sql) if all_sql else ""
    # 图表类型
    chart_type_text = ", ".join(all_chart_types) if all_chart_types else "无"
    # 图形格式显示：包含图表类型和图片文件名
    if all_images:
        image_text = ", ".join(all_images)
        chart_display = f"{chart_type_text} ({image_text})" if chart_type_text != "无" else image_text
    else:
        chart_display = chart_type_text

    return {
        "sql": sql_text,
        "chart_type": chart_type_text,
        "chart_display": chart_display,
        "images": all_images,
        "references": all_references,
    }


def save_result_xlsx(results: list, save_path: str):
    """保存结果为Excel文件（附件7表6格式，含references）"""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "result_3"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D4D4D4'),
        right=Side(style='thin', color='D4D4D4'),
        top=Side(style='thin', color='D4D4D4'),
        bottom=Side(style='thin', color='D4D4D4'),
    )

    headers = ["编号", "问题", "SQL查询语法", "图形格式", "回答"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    wrap_alignment = Alignment(vertical='top', wrap_text=True)
    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r["id"]).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=row_idx, column=2, value=json.dumps(r["questions"], ensure_ascii=False)).alignment = wrap_alignment
        ws.cell(row=row_idx, column=3, value=r.get("sql", "")).alignment = wrap_alignment
        ws.cell(row=row_idx, column=4, value=r.get("chart_display", r.get("chart_type", "无"))).alignment = wrap_alignment
        ws.cell(row=row_idx, column=5, value=json.dumps(r["answers"], ensure_ascii=False)).alignment = wrap_alignment

        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).border = thin_border

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 65
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 100

    wb.save(save_path)
    logger.info(f"结果已保存: {save_path}")


async def main():
    start_time = time.time()

    print("=" * 70)
    print("  任务三：增强智能问数助手 - 回答附件6的问题")
    print("  开始时间:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    print("=" * 70)

    config = AppConfig.load()
    db = DatabaseManager(config.db_path)
    llm = LLMClient(config)
    embedding = EmbeddingManager(config)

    # 检查数据库
    row_count = db.get_table_row_count("income_sheet")
    if row_count == 0:
        print("错误：数据库为空，请先运行 task1/run_task1.py")
        return
    print(f"\n数据库状态: income_sheet 有 {row_count} 条记录")

    # 构建知识库
    kb = build_knowledge_base(db, embedding, SAMPLE_DATA_DIR)

    # 初始化增强版Agent
    results_dir = str(RESULTS_DIR)
    os.makedirs(results_dir, exist_ok=True)

    agent = SmartQAAgent(db=db, llm=llm, knowledge_base=kb, results_dir=results_dir)
    print("\n增强版智能问答Agent初始化完成\n")

    # 加载问题
    questions = load_questions(SAMPLE_DATA_DIR)
    if not questions:
        print("未找到问题文件，使用示例问题")
        questions = [
            {"id": "B2001", "type": "多意图", "questions": [{"Q": "2024年利润最高的top10企业是哪些？这些企业的利润、销售额年同比是多少？年同比上涨幅度最大的是哪家企业？"}]},
            {"id": "B2002", "type": "意图模糊", "questions": [{"Q": "国家医保目录新增的中药产品有哪些"}]},
            {"id": "B2003", "type": "归因分析", "questions": [{"Q": "华润三九近三年的主营业务收入情况做可视化绘图"}, {"Q": "主营业务收入上升的原因是什么"}]},
        ]

    print(f"共 {len(questions)} 道问题待处理\n")

    # 并发处理（每组多轮对话内部顺序执行，不同组之间并发）
    CONCURRENCY = 10
    sem = asyncio.Semaphore(CONCURRENCY)
    all_results = [None] * len(questions)
    done_count = 0
    fail_count = 0
    phase_start = time.time()

    async def _process_one(idx, q):
        nonlocal done_count, fail_count
        async with sem:
            try:
                answers = await agent.process_conversation(
                    questions=q["questions"],
                    enhanced_mode=True,
                    question_id=q["id"],
                )
                extracted = _extract_results_from_answers(answers)
                all_results[idx] = {
                    "id": q["id"],
                    "questions": q["questions"],
                    "answers": answers,
                    "sql": extracted["sql"],
                    "chart_type": extracted["chart_type"],
                    "chart_display": extracted["chart_display"],
                    "images": extracted["images"],
                }
            except Exception as e:
                logger.error(f"处理 {q['id']} 失败: {e}")
                all_results[idx] = {
                    "id": q["id"],
                    "questions": q["questions"],
                    "answers": [{"Q": q["questions"][0]["Q"], "A": {"content": f"处理失败: {e}"}}],
                    "sql": "", "chart_type": "无", "chart_display": "无", "images": [],
                }
                fail_count += 1
            finally:
                done_count += 1
                elapsed = time.time() - phase_start
                rate = done_count / elapsed if elapsed > 0 else 0
                eta = (len(questions) - done_count) / rate if rate > 0 else 0
                if done_count % 5 == 0 or done_count == len(questions):
                    print(f"  进度: {done_count}/{len(questions)} | "
                          f"失败{fail_count} | {rate:.1f}题/秒 | "
                          f"剩余~{eta:.0f}秒", flush=True)

    tasks = [_process_one(idx, q) for idx, q in enumerate(questions)]
    await asyncio.gather(*tasks)

    # 过滤None
    all_results = [r for r in all_results if r is not None]

    # 保存结果
    result_path = str(RESULTS_DIR / "result_3.xlsx")
    save_result_xlsx(all_results, result_path)

    json_path = str(RESULTS_DIR / "result_3.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)

    # 打印统计
    total_sql = sum(1 for r in all_results if r["sql"])
    total_charts = sum(1 for r in all_results if r["chart_type"] != "无")
    total_images = sum(len(r.get("images", [])) for r in all_results)
    elapsed = time.time() - start_time
    print(f"\n{'=' * 70}")
    print(f"  任务三完成！耗时: {elapsed:.1f} 秒")
    print(f"  结果文件: {result_path}")
    print(f"  SQL语句: {total_sql}/{len(all_results)} 道题")
    print(f"  图表生成: {total_charts} 个 ({total_images} 张图片)")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    os.makedirs(str(PROJECT_ROOT / "logs"), exist_ok=True)
    os.makedirs(str(RESULTS_DIR), exist_ok=True)
    asyncio.run(main())
